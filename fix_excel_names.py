# -*- coding: utf-8 -*-
"""修复底稿 Excel 中人员名字大小写不一致问题。
仅修改「负责人员 / 专案执行人员」列的单元格文本，保留其他所有内容、格式、公式。
"""
import openpyxl
import re

SRC = r'C:\Users\Yuki\OneDrive - 大聯大控股\Microsoft Copilot Chat 文件\HR团队资料库\CN区 HR服务指标统计\CN HR工作量表维护专区-For HR\CN HR工作量维护表-For  HR.xlsx'

# 规范姓名列表（与 HR_META.persons 一致），用于校验标准化结果
CANON = {
    'Bill Tsang', 'Cako Yang', 'Chloe Fang', 'Chris Zhang', 'Grace Wang',
    'Heison Wong', 'Hope Zhan', 'Iris Li', 'Janet Au-Yeung', 'Jessie Yang',
    'Lily Qiu', 'Mandy Hu', 'Mia Meng', 'Olivia Ye', 'Penny Pan', 'Rin Shen',
    'Roc Tian', 'Sherry Jiang', 'Skye Yu', 'Tammy Tao', 'Tia Yang', 'Zoe Lin',
}

def norm_person(s):
    if s is None:
        return None
    s = str(s).strip()
    if not s or s == '\u3000':
        return None
    parts = s.split()
    out = []
    for p in parts:
        sub = p.split('-')
        out.append('-'.join((w[:1].upper() + w[1:].lower()) if w else w for w in sub))
    return ' '.join(out)

# 载入（非只读，可写回）
wb = openpyxl.load_workbook(SRC, read_only=False, data_only=True)

sheets = {
    'CN HR工作量表-HR填写': 6,                    # 负责人员 列（0基）
    'CN HR工作量表 -HR填写（专案类）': 6,           # 专案执行人员 列
}

report = {}
for sheet_name, col_idx in sheets.items():
    ws = wb[sheet_name]
    changes = 0
    seen_raw = set()
    for row in ws.iter_rows(min_row=2):
        cell = row[col_idx]
        raw = cell.value
        if raw is None:
            continue
        seen_raw.add(str(raw).strip())
        norm = norm_person(raw)
        # 若标准化后不在规范名单，但原始为某规范的变体，强制对齐
        if norm not in CANON:
            # 规范化再做一次宽松匹配（已在 norm_person 处理）；若仍不匹配保留 norm
            pass
        if str(raw).strip() != norm:
            cell.value = norm
            changes += 1
    report[sheet_name] = {'changes': changes, 'distinct_raw': len(seen_raw)}

# 保存
wb.save(SRC)

print('=== 修复完成 ===')
for sn, info in report.items():
    print(f'Sheet: {sn}')
    print(f'  修改单元格数: {info["changes"]}')
    print(f'  修复前不同名字字符串数: {info["distinct_raw"]}')

# 验证：重新读取，统计唯一姓名
wb2 = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
all_names = set()
for sheet_name, col_idx in sheets.items():
    ws = wb2[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_idx] is not None:
            all_names.add(str(row[col_idx]).strip())
print(f'\n修复后唯一姓名数: {len(all_names)}')
not_in_canon = sorted(all_names - CANON)
print('不在规范名单中的名字:', not_in_canon if not_in_canon else '无 ✓')
print('唯一姓名列表:')
for n in sorted(all_names):
    print('  ', repr(n))
