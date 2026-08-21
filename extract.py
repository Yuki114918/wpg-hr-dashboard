# -*- coding: utf-8 -*-
import openpyxl, json, re, os

SRC = r'C:/Users/Yuki/OneDrive - 大聯大控股/Microsoft Copilot Chat 文件/HR团队资料库/CN区 HR服务指标统计/CN HR工作量表维护专区-For HR/CN HR工作量维护表-For  HR.xlsx'
OUT_DIR = r'C:\Users\Yuki\WorkBuddy\AI 文件处理\WOKE\2026-08-20-15-45-10'
os.makedirs(OUT_DIR, exist_ok=True)

def norm_person(s):
    if s is None: return None
    s = str(s).strip()
    if not s or s == '\u3000': return None
    parts = s.split()
    out = []
    for p in parts:
        sub = p.split('-')
        out.append('-'.join((w[:1].upper() + w[1:].lower()) if w else w for w in sub))
    return ' '.join(out)

def norm_bu(s):
    if s is None: return None
    s = str(s).strip()
    if not s or s == '\u3000': return None
    low = s.lower()
    if low == 'vsell': return 'VSELL'
    if s in ('其它', '其他'): return '其他'
    if low == 'other': return '其他'
    return s

def norm_txt(s):
    if s is None: return None
    s = str(s).replace('\xa0', ' ').replace('\u3000', '').strip()
    return s if s else None

def fmt_ym(v):
    if v is None: return None
    s = str(v).strip()
    if re.fullmatch(r'\d{6}', s):
        return s[:4] + '-' + s[4:6]
    # maybe excel date
    try:
        import datetime
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v.strftime('%Y-%m')
    except Exception:
        pass
    return s

def is_meeting(v):
    if v is None: return False
    s = str(v).strip()
    return s != '' and s != '\u3000'

def num(v):
    if v is None or str(v).strip() == '': return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
records = []

# ---- Daily sheet ----
ws = wb['CN HR工作量表-HR填写']
for row in ws.iter_rows(min_row=2, values_only=True):
    module = norm_txt(row[1])
    person = norm_person(row[6])
    if module is None and person is None:
        continue
    records.append({
        'type': 'daily',
        'module': module or '未分类',
        'workitem': norm_txt(row[2]) or '未分类',
        'detail': norm_txt(row[3]) or '未分类',
        'meeting': is_meeting(row[4]),
        'remark': norm_txt(row[5]),
        'person': person or '未填写',
        'qty': num(row[7]),
        'unit': norm_txt(row[8]),
        'hours': num(row[9]),
        'bu': norm_bu(row[10]),
        'region': norm_txt(row[11]) or '未填写',
        'ym': fmt_ym(row[12]),
        'highvalue': norm_txt(row[13]),
    })

# ---- Project sheet ----
ws = wb['CN HR工作量表 -HR填写（专案类）']
for row in ws.iter_rows(min_row=2, values_only=True):
    module = norm_txt(row[1])
    person = norm_person(row[6])
    if module is None and person is None:
        continue
    records.append({
        'type': 'project',
        'module': module or '未分类',
        'workitem': norm_txt(row[2]) or '未分类',
        'detail': norm_txt(row[3]) or '未分类',
        'meeting': is_meeting(row[4]),
        'remark': norm_txt(row[5]),
        'person': person or '未填写',
        'qty': num(row[7]),
        'unit': norm_txt(row[8]),
        'hours': num(row[9]),
        'bu': None,  # project sheet has no BU column
        'region': norm_txt(row[10]) or '未填写',
        'ym': fmt_ym(row[11]),
        'highvalue': None,
    })

# ---- Build meta ----
persons = sorted({r['person'] for r in records if r['person'] != '未填写'})
modules = sorted({r['module'] for r in records})
regions = sorted({r['region'] for r in records if r['region'] != '未填写'})
bus = sorted({r['bu'] for r in records if r['bu']})
yms = sorted({r['ym'] for r in records if r['ym']})
years = sorted({y[:4] for y in yms})

meta = {
    'persons': persons,
    'modules': modules,
    'regions': regions,
    'bus': bus,
    'yms': yms,
    'years': years,
    'counts': {
        'daily': sum(1 for r in records if r['type'] == 'daily'),
        'project': sum(1 for r in records if r['type'] == 'project'),
        'total': len(records),
    },
    'hours': {
        'daily': round(sum(r['hours'] for r in records if r['type'] == 'daily'), 1),
        'project': round(sum(r['hours'] for r in records if r['type'] == 'project'), 1),
    }
}

# verify normalization fixed dupes
raw_persons_daily = set()
ws2 = wb['CN HR工作量表-HR填写']
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row[6] is not None:
        raw_persons_daily.add(str(row[6]).strip())
raw_persons_proj = set()
ws3 = wb['CN HR工作量表 -HR填写（专案类）']
for row in ws3.iter_rows(min_row=2, values_only=True):
    if row[6] is not None:
        raw_persons_proj.add(str(row[6]).strip())

print('RAW daily person strings:', len(raw_persons_daily), '-> normalized:', len(persons))
print('RAW project person strings:', len(raw_persons_proj))
print('META years:', years)
print('META yms:', yms)
print('META modules:', modules)
print('META regions:', regions)
print('META bus count:', len(bus), bus)
print('META counts:', meta['counts'])
print('META hours:', meta['hours'])
print('Total records:', len(records))

# Write data.js
with open(os.path.join(OUT_DIR, 'hr_data.js'), 'w', encoding='utf-8') as f:
    f.write('window.HR_META = ' + json.dumps(meta, ensure_ascii=False) + ';\n')
    f.write('window.HR_RECORDS = ' + json.dumps(records, ensure_ascii=False) + ';\n')

print('Wrote hr_data.js')
